#!/usr/bin/env python3
"""
Convert docx/pdf files to Markdown format.
Extracts images, tables, and preserves text formatting (bold, italic).
"""

import sys
import subprocess
import re
from pathlib import Path
from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph
from zipfile import ZipFile

# For PDF support (optional)
try:
    import pypdf
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# For PDF table extraction (optional)
try:
    import pdfplumber
    PDF_TABLE_SUPPORT = True
except ImportError:
    PDF_TABLE_SUPPORT = False


def clean_markdown_text(text: str) -> str:
    """Clean up markdown text by fixing bold/italic marker spacing."""
    if not text:
        return text

    text = text.strip()
    # Remove consecutive bold markers: **** -> (empty)
    text = re.sub(r'\*\*\*\*', '', text)
    # Remove bold markers that contain only whitespace: **   ** -> ' '
    text = re.sub(r'\*\*\s+\*\*', ' ', text)
    # Remove italic markers that contain only whitespace: *   * -> ' '
    text = re.sub(r'\*(?!\*)\s+\*(?!\*)', ' ', text)
    # Fix spaces within bold markers: **   text   ** -> **text**
    text = re.sub(r'\*\*\s+', '**', text)
    text = re.sub(r'\s+\*\*', '**', text)
    # Fix spaces within italic markers: *   text   * -> *text*
    text = re.sub(r'\*(?!\*)\s+', '*', text)
    text = re.sub(r'\s+\*(?!\*)', '*', text)

    return text


def guess_extension(data: bytes) -> str:
    """Guess image file extension from magic bytes."""
    if data[:8] == b'\x89PNG\r\n\x1a\n':
        return '.png'
    elif data[:2] == b'\xff\xd8':
        return '.jpeg'
    elif data[:6] in (b'GIF87a', b'GIF89a'):
        return '.gif'
    elif data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return '.webp'
    elif data[:4] == b'BM':
        return '.bmp'
    else:
        return '.png'  # default


def extract_image_number(image_path: str) -> int:
    """Extract the numeric suffix from image path (e.g., 'image10.png' -> 10)."""
    import re
    match = re.search(r'image(\d+)', image_path)
    if match:
        return int(match.group(1))
    return 0


def fix_corrupted_docx(docx_path: Path) -> Path:
    """Fix corrupted docx files by removing invalid relationships (e.g., Target='../NULL').
    Returns path to the fixed file (temporary file that caller should clean up)."""
    import tempfile
    import os

    # Create a temporary file
    temp_fd, temp_path = tempfile.mkstemp(suffix='.docx')
    temp_path = Path(temp_path)

    try:
        with ZipFile(docx_path, 'r') as original_zip:
            with ZipFile(temp_path, 'w') as fixed_zip:
                for item in original_zip.infolist():
                    data = original_zip.read(item.filename)

                    # Fix relationships file - remove invalid relationships
                    if item.filename == 'word/_rels/document.xml.rels':
                        # Remove invalid relationships with Target="../NULL" or Target="NULL"
                        data = data.replace(b'Target="../NULL"', b'Target=""')
                        data = data.replace(b'Target="NULL"', b'Target=""')

                    # Write with the original filename
                    fixed_zip.writestr(item.filename, data)

        # Close the file descriptor
        os.close(temp_fd)

        return temp_path

    except Exception as e:
        # Clean up temp file if something goes wrong
        try:
            os.close(temp_fd)
        except:
            pass
        try:
            temp_path.unlink()
        except:
            pass
        raise


def extract_text_from_xml(docx_path: Path) -> tuple[str, int, int]:
    """Extract text with formatting directly from docx XML.
    This is more accurate than mammoth for corrupted files with formatting issues.

    Returns:
        tuple: (markdown_text, image_count, table_count)
    """
    import xml.etree.ElementTree as ET

    lines = []
    image_counter = 0
    table_counter = 0

    try:
        with ZipFile(docx_path, 'r') as docx_zip:
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
                  'pic': 'http://schemas.openxmlformats.org/drawingml/2006/picture',
                  'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}

            # Parse relationships to check for valid image references
            valid_image_rids = set()
            try:
                with docx_zip.open('word/_rels/document.xml.rels') as rels_xml:
                    rels_tree = ET.parse(rels_xml)
                    for rel in rels_tree.getroot():
                        rel_id = rel.get('Id')
                        target = rel.get('Target')
                        # Only count relationships that point to actual media files
                        if target and target.startswith('media/'):
                            valid_image_rids.add(rel_id)
            except:
                pass

            # First, parse styles.xml to get character style bold information
            bold_char_styles = set()
            try:
                with docx_zip.open('word/styles.xml') as styles_xml:
                    styles_tree = ET.parse(styles_xml)
                    styles_root = styles_tree.getroot()

                    # Find all character styles that are bold (explicit type="character")
                    for style in styles_root.findall('.//w:style[@w:type="character"]', ns):
                        rpr = style.find('w:rPr', ns)
                        if rpr is not None:
                            b = rpr.find('w:b', ns)
                            if b is not None:
                                style_id = style.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}styleId')
                                if style_id:
                                    bold_char_styles.add(style_id)

                    # Also check all styles (not just character type) for bold default
                    # Some styles may not have explicit type attribute
                    for style in styles_root.findall('.//w:style', ns):
                        style_id = style.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}styleId')
                        if not style_id:
                            continue

                        # Check if this style has bold formatting
                        rpr = style.find('w:rPr', ns)
                        if rpr is not None:
                            b = rpr.find('w:b', ns)
                            if b is not None:
                                # Check if w:b/@w:val is not "false" or 0
                                val = b.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '1')
                                if val != '0' and val != 'false':
                                    bold_char_styles.add(style_id)
            except:
                # No styles.xml or error reading it
                pass

            # Now parse document.xml
            with docx_zip.open('word/document.xml') as doc_xml:
                tree = ET.parse(doc_xml)
                root = tree.getroot()

                # Find the body element
                body = root.find('.//w:body', ns)
                if body is None:
                    body = root

                # Collect all table cell elements to identify nested paragraphs
                table_cells = set()
                for tc in body.findall('.//w:tc', ns):
                    for p in tc.findall('.//w:p', ns):
                        table_cells.add(id(p))

                # Track which paragraphs we've already processed (to skip table-cell paragraphs)
                processed_paragraphs = set()

                # Track which tables we've processed (to skip nested tables)
                processed_tables = set()

                # Process body elements in order to maintain paragraph/table ordering
                for elem in body:
                    # Check if this is a paragraph (w:p)
                    if elem.tag.endswith('p') or elem.tag == '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p':
                        # Skip paragraphs inside table cells
                        if id(elem) in table_cells:
                            continue

                        # Skip already processed paragraphs
                        if id(elem) in processed_paragraphs:
                            continue
                        processed_paragraphs.add(id(elem))

                        # Check if this paragraph contains images
                        # Count only w:drawing elements with valid image references
                        images_in_para = 0
                        for drawing in elem.findall('.//w:drawing', ns):
                            # Check if this drawing has a valid blip with r:embed
                            blip = drawing.find('.//a:blip', ns)
                            if blip is not None:
                                embed = blip.get(f'{{{ns["r"]}}}embed')
                                if embed and embed in valid_image_rids:
                                    images_in_para += 1

                        # Get runs and their formatting
                        runs = elem.findall('.//w:r', ns)

                        if not runs and images_in_para == 0:
                            continue

                        # Collect (text, is_bold) tuples, preserving all text including whitespace
                        run_data = []
                        non_empty_runs = []  # Runs that have non-whitespace content

                        for run in runs:
                            # Get text from this run - preserve original text
                            texts = []
                            for t in run.findall('.//w:t', ns):
                                if t.text:
                                    texts.append(t.text)

                            run_text = ''.join(texts)
                            if not run_text:
                                continue  # Skip completely empty runs

                            # Check if this run is bold (from direct formatting or character style)
                            rpr = run.find('w:rPr', ns)
                            is_bold = False

                            if rpr is not None:
                                # Check direct bold formatting
                                b = rpr.find('w:b', ns)
                                is_bold = b is not None

                                # If not directly bold, check if it has a bold character style
                                if not is_bold:
                                    rstyle = rpr.find('w:rStyle', ns)
                                    if rstyle is not None:
                                        style_val = rstyle.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                                        if style_val in bold_char_styles:
                                            is_bold = True

                            # For whitespace-only runs, treat as non-bold (don't add ** markers)
                            has_content = run_text.strip() != ''
                            effective_bold = is_bold and has_content

                            run_data.append((run_text, effective_bold))

                            # Track runs with non-whitespace content for bold detection
                            if has_content:
                                non_empty_runs.append(is_bold)

                        # Handle image-only paragraphs (no text runs)
                        if not run_data and images_in_para > 0:
                            para_text = ''
                        elif not run_data:
                            continue  # Skip paragraphs with no content at all
                        else:
                            # Check if all runs with actual content are bold
                            all_bold = non_empty_runs and all(is_bold for is_bold in non_empty_runs)
                            any_bold = any(is_bold for _, is_bold in run_data)

                            # Build the paragraph text
                            if all_bold and any_bold:
                                # All runs with content are bold - combine and wrap once
                                combined_text = ''.join(text for text, _ in run_data)
                                # Strip leading/trailing whitespace from the content inside bold markers
                                combined_text = combined_text.strip()
                                para_text = f'**{combined_text}**'
                            else:
                                # Mixed formatting or no bold - merge consecutive bold runs
                                para_parts = []
                                i = 0
                                while i < len(run_data):
                                    text, is_bold = run_data[i]

                                    if is_bold:
                                        # Start of a bold sequence - find all consecutive bold runs
                                        bold_text_parts = [text]
                                        j = i + 1
                                        while j < len(run_data) and run_data[j][1]:
                                            bold_text_parts.append(run_data[j][0])
                                            j += 1
                                        # Combine consecutive bold runs and strip leading/trailing whitespace
                                        bold_content = ''.join(bold_text_parts).strip()
                                        para_parts.append(f'**{bold_content}**')
                                        i = j
                                    else:
                                        para_parts.append(text)
                                        i += 1

                                # Strip leading/trailing whitespace from final paragraph
                                para_text = ''.join(para_parts).strip()

                        # Append image markers if this paragraph has images
                        if images_in_para > 0:
                            markers = []
                            for _ in range(images_in_para):
                                image_counter += 1
                                markers.append(f'(image-{image_counter})')
                            if para_text:
                                para_text = f"{para_text} {' '.join(markers)}"
                            else:
                                para_text = ' '.join(markers)

                        if para_text:
                            lines.append(para_text)

                    # Check if this is a table (w:tbl)
                    elif elem.tag.endswith('tbl') or elem.tag == '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tbl':
                        # Skip nested tables (tables inside table cells)
                        # Find parent to check if we're inside a table cell
                        parent_map = {c: p for p in body.iter() for c in p}
                        current = elem
                        is_nested = False
                        while current in parent_map:
                            parent = parent_map[current]
                            if parent.tag.endswith('tc'):
                                is_nested = True
                                break
                            current = parent

                        if is_nested:
                            continue

                        # Skip already processed tables
                        if id(elem) in processed_tables:
                            continue
                        processed_tables.add(id(elem))

                        # Add table marker
                        table_counter += 1
                        lines.append(f'(table-{table_counter})')

    except Exception as e:
        print(f"  Warning: Could not extract text from XML: {e}")
        return "", 0, 0

    # Clean up lines using shared function
    cleaned_lines = [clean_markdown_text(line) for line in lines]
    return '\n\n'.join(cleaned_lines), image_counter, table_counter


def extract_text_with_mammoth(docx_path: Path, fixed_path: Path = None) -> tuple[str, int, int]:
    """Use mammoth library to extract text from corrupted docx files with formatting.

    NOTE: For files with formatting issues, prefer extract_text_from_xml() instead.

    Args:
        docx_path: Path to the original docx file
        fixed_path: Path to the fixed docx file (if fix_corrupted_docx was used)

    Returns:
        tuple: (markdown_text, image_count, table_count)
    """
    # For now, use the XML-based extraction for more accurate formatting
    return extract_text_from_xml(docx_path)


def extract_tables_from_xml(docx_path: Path) -> list[list[list[str]]]:
    """Extract tables directly from docx document.xml."""
    import xml.etree.ElementTree as ET

    tables_data = []

    try:
        with ZipFile(docx_path, 'r') as docx_zip:
            # Read document.xml
            with docx_zip.open('word/document.xml') as doc_xml:
                tree = ET.parse(doc_xml)
                root = tree.getroot()

                # Define namespace
                ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}

                # Find all tables
                for table in root.findall('.//w:tbl', ns):
                    table_data = []
                    for row in table.findall('.//w:tr', ns):
                        row_data = []
                        for cell in row.findall('.//w:tc', ns):
                            # Get all text runs in the cell
                            text_parts = []
                            for paragraph in cell.findall('.//w:p', ns):
                                for text in paragraph.findall('.//w:t', ns):
                                    if text.text:
                                        text_parts.append(text.text)
                            cell_text = ''.join(text_parts).strip()
                            row_data.append(cell_text)
                        table_data.append(row_data)
                    if table_data:
                        tables_data.append(table_data)

    except Exception as e:
        print(f"  Warning: Could not extract tables from XML: {e}")

    return tables_data


class DocxToMarkdown:
    """Convert DOCX to Markdown with image and table extraction."""

    def __init__(self, docx_path: str) -> None:
        self.docx_path = Path(docx_path)
        self._temp_file = None  # Track temporary file for cleanup
        self._use_mammoth = False  # Flag for using mammoth fallback
        self.doc = None

        # Try to load the document with various methods
        try:
            self.doc = Document(docx_path)
        except KeyError as e:
            if 'NULL' in str(e):
                print("  Detected corrupted docx file, attempting to fix...")
                try:
                    self._temp_file = fix_corrupted_docx(self.docx_path)
                    self.doc = Document(str(self._temp_file))
                    print("  Fixed and loaded successfully.")
                except Exception:
                    print("  Fix failed, using mammoth library as fallback...")
                    self._use_mammoth = True
            else:
                print("  Using mammoth library as fallback...")
                self._use_mammoth = True
        except Exception:
            print("  Using mammoth library as fallback...")
            self._use_mammoth = True

        self.images_dir = self.docx_path.parent / "images"
        self.image_counter = 0
        self.markdown_lines: list[str] = []
        self.first_line_processed = False
        # Get directory name for image naming (e.g., "news-1" from "/path/to/news-1/file.docx")
        # If parent is current directory (name is empty), use the actual directory name
        self.dir_name = self.docx_path.parent.name or self.docx_path.parent.resolve().name
        self.extracted_images: list[str] = []
        self.tables_data: list[list[list[str]]] = []  # Store tables for xlsx export

    def __del__(self):
        """Clean up temporary file if it exists."""
        if self._temp_file and self._temp_file.exists():
            try:
                self._temp_file.unlink()
            except:
                pass

    def iter_block_items(self, parent):
        """Yield each paragraph and table child within *parent*."""
        if isinstance(parent, _Cell):
            el_func = parent._tc
        elif isinstance(parent, DocumentType):
            el_func = parent.element.body
        else:
            raise ValueError("Unsupported parent type")

        for element in el_func.iterchildren():
            if isinstance(element, CT_P):
                yield Paragraph(element, parent)
            elif isinstance(element, CT_Tbl):
                yield Table(element, parent)

    def extract_all_images(self) -> None:
        """Extract all images from the docx file and store paths."""
        self.images_dir.mkdir(exist_ok=True)

        with ZipFile(self.docx_path) as docx_zip:
            # Find all image files, excluding the directory entry itself
            # Sort by numeric suffix to maintain correct order (image1, image2, ... image10, image11)
            image_files = sorted([f for f in docx_zip.namelist()
                                if f.startswith('word/media/') and f != 'word/media/'],
                               key=extract_image_number)

            for image_path in image_files:
                try:
                    with docx_zip.open(image_path) as source:
                        data = source.read()

                    # Skip if empty or too small
                    if len(data) < 100:
                        print(f"  Skipping {image_path} (too small: {len(data)} bytes)")
                        continue

                    # Get file extension
                    ext = guess_extension(data)

                    # Generate image name: directory-name-image-sequence
                    self.image_counter += 1
                    image_name = f"{self.dir_name}-image-{self.image_counter}{ext}"
                    target_path = self.images_dir / image_name

                    # Save image
                    with open(target_path, 'wb') as target:
                        target.write(data)

                    image_path_relative = f"images/{image_name}"
                    self.extracted_images.append(image_path_relative)
                    print(f"  Saved: {image_name}")

                except Exception as e:
                    print(f"  Error extracting {image_path}: {e}")
                    continue

    def extract_tables_to_xlsx(self) -> None:
        """Extract all tables and save to xlsx files with basic styling."""
        if not self.tables_data:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            print("  Warning: openpyxl not installed. Skip exporting tables to xlsx.")
            print("  Install with: pip install openpyxl")
            return

        tables_dir = self.docx_path.parent / "tables"
        tables_dir.mkdir(exist_ok=True)

        # Define styles
        font_normal = Font(name='Arial', size=12)
        font_bold = Font(name='Arial', size=12, bold=True)
        alignment_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for idx, table_data in enumerate(self.tables_data, 1):
            wb = openpyxl.Workbook()
            ws = wb.active

            for row_idx, row_data in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                    # Set font and alignment
                    cell.font = font_normal
                    cell.alignment = alignment_center
                    cell.border = thin_border

                    # Header row (first row): bold
                    if row_idx == 1:
                        cell.font = font_bold

            # Auto-adjust column widths
            for col in ws.columns:
                col_letter = col[0].column_letter
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            value_str = str(cell.value)
                            chinese_chars = sum(1 for c in value_str if '\u4e00' <= c <= '\u9fff')
                            other_chars = len(value_str) - chinese_chars
                            length = chinese_chars * 2 + other_chars
                            max_length = max(max_length, length)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            xlsx_path = tables_dir / f"{self.dir_name}-table-{idx}.xlsx"
            wb.save(xlsx_path)
            print(f"  Saved table: {xlsx_path}")

    def table_to_list(self, table: Table) -> list[list[str]]:
        """Convert a docx table to a list of lists."""
        table_data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                row_data.append(cell.text.strip())
            table_data.append(row_data)
        return table_data

    def count_images_in_paragraph(self, paragraph) -> int:
        """Count how many images are in this paragraph."""
        count = 0
        for run in paragraph.runs:
            if 'pic:pic' in run._element.xml:
                count += 1
        return count

    def format_run(self, run) -> str:
        """Format a text run with markdown styling."""
        text = run.text
        if not text:
            return ""

        # Check for styles using multiple methods
        is_bold = False
        is_italic = False

        # Method 1: Use python-docx built-in properties
        try:
            is_bold = run.bold if run.bold is not None else False
            is_italic = run.italic if run.italic is not None else False
        except:
            pass

        # Method 2: Direct XML parsing for more reliable detection
        if not is_bold or not is_italic:
            try:
                import xml.etree.ElementTree as ET
                ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
                rpr = run._element.find('.//w:rPr', ns)
                if rpr is not None:
                    # Check bold
                    b = rpr.find('w:b', ns)
                    if b is not None:
                        val = b.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '1')
                        is_bold = is_bold or (val != '0' and val != 'false')
                    # Check italic
                    i = rpr.find('w:i', ns)
                    if i is not None:
                        val = i.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '1')
                        is_italic = is_italic or (val != '0' and val != 'false')
            except:
                pass

        # Apply markdown formatting
        if is_bold and is_italic:
            return f"***{text}***"
        elif is_bold:
            return f"**{text}**"
        elif is_italic:
            return f"*{text}*"
        else:
            return text

    def paragraph_to_markdown(self, paragraph, image_index: int) -> tuple[str, int]:
        """Convert a paragraph to markdown. Returns (md_text, images_used_count)."""
        text = paragraph.text.strip()
        if not text:
            # Check if this paragraph contains only images
            images_in_para = self.count_images_in_paragraph(paragraph)
            if images_in_para > 0:
                # Insert image markers at this position
                markers = []
                for i in range(images_in_para):
                    self.image_counter += 1
                    markers.append(f"(image-{self.image_counter})")
                return ' '.join(markers), images_in_para
            return "", 0

        # Format runs with styling
        formatted_text = ''.join(self.format_run(run) for run in paragraph.runs)

        # Count images in this paragraph
        images_in_para = self.count_images_in_paragraph(paragraph)
        images_used = images_in_para

        # If paragraph has images, append image markers after the text
        if images_in_para > 0:
            markers = []
            for i in range(images_in_para):
                self.image_counter += 1
                markers.append(f"(image-{self.image_counter})")
            return f"{formatted_text} {' '.join(markers)}", images_in_para

        return formatted_text, images_used

    def convert(self) -> str:
        """Convert the docx to markdown."""
        # Check if using mammoth fallback mode
        if self._use_mammoth:
            print("Using mammoth library for text extraction...")
            text, extracted_image_count, extracted_table_count = extract_text_with_mammoth(
                self.docx_path, self._temp_file)

            # Extract images (the text already has image markers embedded)
            print("Extracting images...")
            self.extract_all_images()
            print(f"Extracted {len(self.extracted_images)} images\n")

            # Extract tables from XML (the text already has table markers embedded)
            print("Attempting to extract tables from XML...")
            self.tables_data = extract_tables_from_xml(self.docx_path)
            if self.tables_data:
                print(f"Found {len(self.tables_data)} tables\n")
            else:
                print("No tables found\n")

            # Process text into lines - text already has (image-N) and (table-N) markers
            import re
            # Split by paragraph separator (double newlines)
            lines = text.split('\n\n')
            for line in lines:
                line = line.strip()
                if line:
                    if not self.first_line_processed:
                        # For heading, remove image/table markers from the title line
                        line_for_heading = re.sub(r'\s*\(image-\d+\)', '', line)
                        line_for_heading = re.sub(r'\s*\(table-\d+\)', '', line_for_heading)
                        line_for_heading = line_for_heading.strip()
                        if line_for_heading:
                            line = f"# {line_for_heading}"
                            # Remove all bold/italic markers from heading
                            line = line.replace('**', '').replace('*', '')
                            self.first_line_processed = True
                        else:
                            # Title was only markers, use the original line
                            line = f"# {line}"
                            line = line.replace('**', '').replace('*', '')
                            self.first_line_processed = True
                    else:
                        # For non-heading lines, keep the markers as-is
                        pass

                    self.markdown_lines.append(line)

            # Export tables to xlsx
            if self.tables_data:
                print(f"Extracting {len(self.tables_data)} tables...")
                self.extract_tables_to_xlsx()

            return '\n\n'.join(self.markdown_lines)

        # Standard mode: First, extract all images
        print("Extracting images...")
        self.extract_all_images()
        print(f"Extracted {len(self.extracted_images)} images\n")

        # Reset image counter for markdown markers (will increment during paragraph processing)
        self.image_counter = 0
        # Process document
        self.table_counter = 0  # Track table count for markers

        for block in self.iter_block_items(self.doc):
            if isinstance(block, Paragraph):
                md_text, images_used = self.paragraph_to_markdown(block, 0)

                if md_text:
                    # Only convert first non-empty line to heading
                    if not self.first_line_processed:
                        md_text = f"# {md_text}"
                        # Remove all bold/italic markers from heading
                        md_text = md_text.replace('**', '').replace('*', '')
                        self.first_line_processed = True

                    # Clean up using shared function
                    md_text = clean_markdown_text(md_text)
                    self.markdown_lines.append(md_text)

            elif isinstance(block, Table):
                # Extract table data for xlsx export
                table_data = self.table_to_list(block)
                self.tables_data.append(table_data)

                # Add table marker at the table position
                self.table_counter += 1
                self.markdown_lines.append(f"(table-{self.table_counter})")

        # Export tables to xlsx
        if self.tables_data:
            print(f"Extracting {len(self.tables_data)} tables...")
            self.extract_tables_to_xlsx()

        return '\n\n'.join(self.markdown_lines)

    def save(self) -> Path:
        """Save the markdown file."""
        md_path = self.docx_path.with_suffix('.md')
        markdown_content = self.convert()

        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)

        return md_path


class PdfToMarkdown:
    """Convert PDF to Markdown with text and table extraction."""

    def __init__(self, pdf_path: str) -> None:
        self.pdf_path = Path(pdf_path)
        self.tables_data: list[list[list[str]]] = []
        self.table_counter = 0
        self.images_dir = self.pdf_path.parent / "images"

    def extract_tables_with_pdfplumber(self) -> None:
        """Extract tables from PDF using pdfplumber."""
        if not PDF_TABLE_SUPPORT:
            return

        try:
            import pdfplumber
        except ImportError:
            return

        print("  Extracting tables with pdfplumber...")
        tables_dir = self.pdf_path.parent / "tables"
        tables_dir.mkdir(exist_ok=True)

        # Get directory name for file naming
        dir_name = self.pdf_path.parent.name

        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            # Filter out empty rows and convert to list of lists
                            table_data = []
                            for row in tables[0]:
                                if row and any(cell is not None and str(cell).strip() for cell in row):
                                    table_data.append([str(cell).strip() if cell is not None else "" for cell in row])
                            if table_data:
                                self.tables_data.append(table_data)

        except Exception as e:
            print(f"  Warning: Could not extract tables with pdfplumber: {e}")

        # Export tables to xlsx
        if self.tables_data:
            self._export_tables_to_xlsx(tables_dir, dir_name)

    def _export_tables_to_xlsx(self, tables_dir: Path, dir_name: str) -> None:
        """Export extracted tables to xlsx files."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            print("  Warning: openpyxl not installed. Skip exporting tables to xlsx.")
            return

        # Define styles
        font_normal = Font(name='Arial', size=12)
        font_bold = Font(name='Arial', size=12, bold=True)
        alignment_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for idx, table_data in enumerate(self.tables_data, 1):
            wb = openpyxl.Workbook()
            ws = wb.active

            for row_idx, row_data in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.font = font_normal
                    cell.alignment = alignment_center
                    cell.border = thin_border

                    # Header row: bold
                    if row_idx == 1:
                        cell.font = font_bold

            # Auto-adjust column widths
            for col in ws.columns:
                col_letter = col[0].column_letter
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            value_str = str(cell.value)
                            chinese_chars = sum(1 for c in value_str if '\u4e00' <= c <= '\u9fff')
                            other_chars = len(value_str) - chinese_chars
                            length = chinese_chars * 2 + other_chars
                            max_length = max(max_length, length)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            xlsx_path = tables_dir / f"{dir_name}-table-{idx}.xlsx"
            wb.save(xlsx_path)
            print(f"  Saved table: {xlsx_path}")

    def convert(self) -> str:
        """Extract text from PDF with table position markers."""
        if not PDF_SUPPORT:
            raise ImportError("pypdf is required for PDF support. Install with: pip install pypdf")

        # First, extract tables to know their positions
        self.extract_tables_with_pdfplumber()

        text_content: list[str] = []

        with open(self.pdf_path, 'rb') as f:
            pdf_reader = pypdf.PdfReader(f)
            for page_num, page in enumerate(pdf_reader.pages, 1):
                text = page.extract_text()
                if text and text.strip():
                    text_content.append(text.strip())

                # Add table marker after each page's text
                # (We add a simple marker; exact positioning would require more complex logic)
                if self.tables_data:
                    # For now, add all table markers at the end
                    pass

        result = '\n\n'.join(text_content)

        # Add table markers at the end (simplified approach)
        # A more sophisticated version would detect table positions within the text
        for i in range(len(self.tables_data)):
            self.table_counter += 1
            result += f"\n\n(table-{self.table_counter})"

        return result

    def save(self) -> Path:
        """Save the markdown file."""
        md_path = self.pdf_path.with_suffix('.md')
        markdown_content = self.convert()

        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)

        return md_path


def convert_markdown_to_html(md_path: Path) -> Path:
    """Convert markdown file to HTML using convert_to_html.py script."""
    script_path = Path(__file__).parent / "convert_to_html.py"

    if not script_path.exists():
        print(f"Error: convert_to_html.py not found at {script_path}")
        sys.exit(1)

    # Run the markdown to html conversion script
    result = subprocess.run(
        [sys.executable, str(script_path), str(md_path)],
        capture_output=True,
        text=True
    )

    if result.returncode != 0:
        print(f"Error converting markdown to HTML:")
        print(result.stderr)
        sys.exit(1)

    # Output path is md_path with .html extension
    html_path = md_path.with_suffix('.html')
    return html_path


def main() -> None:
    """Main entry point with smart mode detection."""
    if len(sys.argv) < 2:
        print("Usage: python convert_to_markdown.py <path-to-file>")
        print("  Input file types:")
        print("    .docx, .pdf  -> Convert to Markdown (.md)")
        print("    .md         -> Convert to HTML (.html)")
        sys.exit(1)

    file_path = sys.argv[1]
    path = Path(file_path)

    if not path.exists():
        print(f"Error: File not found: {file_path}")
        sys.exit(1)

    # Smart mode: determine conversion based on file type
    suffix = path.suffix.lower()

    if suffix == '.md':
        # Markdown to HTML mode
        print(f"Converting Markdown to HTML: {file_path}")
        html_path = convert_markdown_to_html(path)
        print(f"\n✓ HTML saved: {html_path}")

    elif suffix == '.docx':
        # DOCX to Markdown mode
        print(f"Converting DOCX: {file_path}")
        converter = DocxToMarkdown(file_path)
        md_path = converter.save()
        print(f"\n✓ Markdown saved: {md_path}")
        print(f"✓ Images extracted to: {converter.images_dir}")
        if converter.tables_data:
            print(f"✓ Tables saved to: {md_path.parent / 'tables'}")

    elif suffix == '.pdf':
        # PDF to Markdown mode
        print(f"Converting PDF: {file_path}")
        converter = PdfToMarkdown(file_path)
        md_path = converter.save()
        print(f"✓ Markdown saved: {md_path}")

    else:
        print(f"Error: Unsupported file type: {suffix}")
        print("Supported types: .docx, .pdf, .md")
        sys.exit(1)


if __name__ == '__main__':
    main()
